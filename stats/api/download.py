from django.http import HttpResponse

import datetime as dt
import openpyxl as xl
from typing import List


def report(queryset, title: str, headers: List[str], filename: str) -> HttpResponse:
    response = HttpResponse(
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = filename.format(title = title, date = dt.datetime.now().strftime('%Y-%m-%d'))
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook = xl.Workbook()
    worksheet = workbook.active
    worksheet.title = title

    r_num = 1
    for c_num, header in enumerate(headers, 1):
        worksheet.cell(row = r_num, column = c_num, value = header)

    for row in queryset:
        r_num += 1
        
        for c_num, value in enumerate((row.user, row.datetimestamp, row.action, row.input, row.output), 1):
            worksheet.cell(row = r_num, column = c_num, value = str(value))

    workbook.save(response)
    return response